import os
import datetime
import json
import openpyxl


class ExcelManager:
    # 엑셀 파일 경로
    _original_file_path = r"D:\patch.xlsx"

    # JSON 파일 경로
    _json_file_path = r"C:\Users\seungsu\Desktop\materials\hash.json"

    def __init__(self, sheet_name):
        if not os.path.exists(self._original_file_path):
            print("원본 파일이 존재하지 않습니다.")
            print("프로그램을 종료합니다...")
            return

        if not os.path.exists(self._json_file_path):
            print("JSON 파일이 존재하지 않습니다.")
            print("프로그램을 종료합니다...")
            return
        
        with open(self._json_file_path, "r") as fp:
            self.json_dict = json.load(fp)
            print("JSON 파일 로딩 완료")

        default_dir = "D:\patch"

        if not os.path.exists(default_dir):
            print("복사 파일이 저장될 폴더를 만듭니다.")
            os.mkdir(default_dir)

        now = datetime.datetime.now().strftime("%Y%m%d")
        self.excel_file_path = default_dir + "\\patch" + now + ".xlsx"

        with open(self._original_file_path, "rb") as fp:
            with open(self.excel_file_path, "wb") as fp2:
                fp2.write(fp.read())

        self.excel_file = openpyxl.load_workbook(self.excel_file_path)
        self.excel_sheet = self.excel_file.get_sheet_by_name(sheet_name)
        print("엑셀 파일 로딩이 완료되었습니다...")
    

    def get_cell_value(self, row, col: str):
        return self.excel_sheet.cell(row, self._col_to_int(col)).value


    def set_cell_value(self, row, col: str, val):
        self.excel_sheet.cell(row, self._col_to_int(col), val)


    def save_workbook(self):
        print("Excel 작업을 마무리합니다...")
        self.excel_file.save(self.excel_file_path)


    def _col_to_int(self, col: str) -> int:
        val = ord(col) - ord('A') if col.isupper() else ord(col) - ord('a')
        return val + 1
    

    def _work_sheet_props(self):
        sheet = self.excel_sheet
        print(sheet.sheet_properties)
    

if __name__ == "__main__":
    em = ExcelManager("dotnet")
    # em.save_workbook()
    # em._work_sheet_props()