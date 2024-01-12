import os
import datetime
import json
import openpyxl


class ExcelManager:
    # 엑셀 파일 경로
    _original_file_path = r"D:\patch.xlsx"

    # Sheet 이름
    _sheet_name = "dotnet"

    # JSON 파일 경로
    _json_file_path = r"C:\Users\seungsu\Desktop\materials\hash.json"

    # 좌표
    sheet_info = {
        "dotnet": {
            "1607 4.8": {
                "bulletinUrl": [(98, 'D'), (99, 'D'), (100, 'D'), (101, 'D')],
                'PatchDate': (98, 'J'),
                '중요도': (98, 'K'),
                'cve': (100, 'P'),
                '파일명': [(103, 'A'), (104, 'A'), (103, 'P'), (104, 'P')],
                '파일크기': [(103, 'B'), (104, 'B')],
                'MD5': [(103, 'J'), (104, 'J')],
                'VendorUrl': [(103, 'K'), (104, 'K')],
                'Wsus': [(103, 'L'), (104, 'L')],
                "SHA256": [(103, 'U'), (104, 'U')],
                "qnumber": [(98, 'H'), (98, 'I')],
            },

            "1809 3.5 4.7.2": {
                "bulletinUrl": [(107, 'D'), (108, 'D'), (109, 'D'), (110, 'D')],
                'PatchDate': (107, 'J'),
                '중요도': (107, 'K'),
                'cve': (109, 'P'),
                '파일명': [(112, 'A'), (113, 'A'), (112, 'P'), (113, 'P')],
                '파일크기': [(112, 'B'), (113, 'B')],
                'MD5': [(112, 'J'), (113, 'J')],
                'VendorUrl': [(112, 'K'), (113, 'K')],
                'Wsus': [(112, 'L'), (113, 'L')],
                "SHA256": [(112, 'U'), (113, 'U')],
                "qnumber": [(107, 'H'), (107, 'I')],
            }
        },
    }

    def __init__(self):
        if not os.path.exists(self._original_file_path):
            print("원본 파일이 존재하지 않습니다.")
            print("프로그램을 종료합니다...")
            return

        if not os.path.exists(self._json_file_path):
            print("JSON 파일이 존재하지 않습니다.")
            print("프로그램을 종료합니다...")
            return
        
        with open(self._json_file_path, "r") as fp:
            self.json_dict = json.load(fp)
            print("JSON 파일 로딩 완료")

        default_dir = "D:\patch"

        if not os.path.exists(default_dir):
            print("복사 파일이 저장될 폴더를 만듭니다.")
            os.mkdir(default_dir)

        now = datetime.datetime.now().strftime("%Y%m%d")
        self.excel_file_path = default_dir + "\\patch" + now + ".xlsx"

        with open(self._original_file_path, "rb") as fp:
            with open(self.excel_file_path, "wb") as fp2:
                fp2.write(fp.read())

        self.excel_file = openpyxl.load_workbook(self.excel_file_path)
        self.excel_sheet = self.excel_file.get_sheet_by_name(self._sheet_name)
        print("엑셀 파일 로딩이 완료되었습니다...")
    

    def get_cell_value(self, row, col: str):
        return self.excel_sheet.cell(row, self._col_to_int(col)).value


    def set_cell_value(self, row, col: str, val):
        self.excel_sheet.cell(row, self._col_to_int(col), val)


    def save_workbook(self):
        print("Excel 작업을 마무리합니다...")
        self.excel_file.save(self.excel_file_path)


    def test(self):
        sheet = self.sheet_info["dotnet"]

        for version in self.json_dict:
            if version not in sheet:
                print(f"\"{version}\": sheet에 없습니다.")
                continue

            if version not in self.json_dict:
                print(f"\"{version}\": self.json_dict에 없습니다.")
                continue

            excel_location_info = sheet[version]
            os_info_list = self.json_dict[version]

            for os_info in os_info_list:
                # BulletinUrl 
                bulletin_locs = excel_location_info["bulletinUrl"]
                flags = ["en-us", "ja-jp", "ko-kr", "zh-cn"]

                for idx, loc in enumerate(bulletin_locs):
                    qnumber = os_info["qnumber"]
                    val = f"https://support.microsoft.com/{flags[idx]}/help/{qnumber}"
                    self.set_cell_value(loc[0], loc[1], val)
                    print(loc[0], loc[1], val)

                for os_key, os_val in os_info.items():

                    # 위치 정보가 기록되어 있지 않은 경우 continue
                    if os_key not in excel_location_info:
                        print(f"\"{os_key}\": excel_location_info에 없습니다.")
                        continue

                    location = excel_location_info[os_key]

                    if type(location) == tuple:
                        if os_key == "cve" and len(os_val) == 0:
                            print("빈 cve 리스트는 continue 합니다.")
                            continue

                        self.set_cell_value(location[0], location[1], os_val)

                    elif type(location) == list:
                        if os_key == "qnumber":
                            self.set_cell_value(location[0][0], location[0][1], "MS-KB" + os_val)
                            self.set_cell_value(location[1][0], location[1][1], "KB" + os_val)

                        else:
                            for loc in location:
                                self.set_cell_value(loc[0], loc[1], os_val)


    def _col_to_int(self, col: str) -> int:
        val = ord(col) - ord('A') if col.isupper() else ord(col) - ord('a')
        return val + 1
    

    def _work_sheet_props(self):
        sheet = self.excel_sheet
        print(sheet.sheet_properties)
    

if __name__ == "__main__":
    em = ExcelManager()
    # em.save_workbook()
    # em._work_sheet_props()

    em.test()
    em.save_workbook()